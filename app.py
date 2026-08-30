# (unchanged imports and setup)
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_mysqldb import MySQL
from MySQLdb.cursors import DictCursor

import MySQLdb.cursors
import openpyxl
import os
import random
import string
import MySQLdb
import base64
from werkzeug.utils import secure_filename


app = Flask(__name__)
app.secret_key = 'jeevansetu_secret_key'

# MySQL config
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'PASSWD'
app.config['MYSQL_DB'] = 'jeevan'
mysql = MySQL(app)

# Excel paths
base_path = os.getcwd()
patient_excel = os.path.join(base_path, 'patients_data.xlsx')
doctor_excel = os.path.join(base_path, 'doctors_data.xlsx')
admin_excel = os.path.join(base_path, 'admins_data.xlsx')

def create_excel(path, headers):
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(path)

create_excel(patient_excel, ['ID','User ID', 'Full Name', 'Aadhaar', 'Email', 'Password', 'Phone', 'DOB', 'Address'])
create_excel(doctor_excel, ['ID','User ID', 'Full Name', 'Aadhaar', 'Email', 'Password', 'Phone', 'Specialization', 'Hospital'])
create_excel(admin_excel, ['ID','User ID', 'Full Name', 'Aadhaar', 'Email', 'Password', 'Phone', 'Hospital'])

def get_next_id(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return ws.max_row

def generate_user_id(role_prefix):
    return f"{role_prefix}{''.join(random.choices(string.ascii_uppercase + string.digits, k=7))}"

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/about')
def about():
    return render_template('about.html')
@app.route('/labreports', methods=['GET', 'POST'])
def labreports():
    # Check admin session
    if 'loggedin' not in session or session.get('role') != 'admin':
        return redirect('/login')

    if request.method == 'POST':
        patient_id = request.form.get('patient_id')
        patient_name = request.form.get('patient_name')
        file = request.files.get('file')
        hospital_name = session['user']['hospital_name']

        if not patient_id or not patient_name or not file:
            flash('All fields are required!', 'danger')
        else:
            try:
                filename = secure_filename(file.filename)
                file_data = file.read()
                file_type = file.content_type
                report_number = 'RPT' + ''.join(random.choices(string.digits, k=6))

                cursor = mysql.connection.cursor()
                cursor.execute("""
                    INSERT INTO lab_reports (
                        report_number, patient_id, hospital_name,
                        file_data, filename, file_type, patient_name
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (report_number, patient_id, hospital_name, file_data, filename, file_type, patient_name))
                mysql.connection.commit()
                cursor.close()

                flash('Lab report uploaded successfully.', 'success')
                return redirect('/labreports')
            except Exception as e:
                flash('An error occurred while uploading: ' + str(e), 'danger')

    return render_template('labreports.html')

    
@app.route('/feedback')
def feedback():
    return render_template('feedback.html')

@app.route('/register', methods=['GET', 'POST']) 
def register():
    cursor = mysql.connection.cursor()
    if request.method == 'POST':
        role = request.form['role']
        full_name = request.form['full_name']
        aadhaar_card = request.form['aadhaar_card']
        email = request.form['email']
        password = request.form['password']
        phone_number = request.form['phone_number']
        gender = request.form['gender']  # New field

        if role == 'patient':
            user_id = generate_user_id("PAT")
            dob = request.form['dob']
            address = request.form['address']
            cursor.execute("SELECT * FROM patients WHERE aadhaar_card = %s OR email = %s", (aadhaar_card, email))
            if cursor.fetchone():
                flash("Patient already registered.", "danger")
                return redirect(url_for('register'))

            cursor.execute('''INSERT INTO patients (user_id, full_name, aadhaar_card, email, password, phone_number, gender, dob, address) 
                              VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                           (user_id, full_name, aadhaar_card, email, password, phone_number, gender, dob, address))
            mysql.connection.commit()
            id_val = get_next_id(patient_excel)
            wb = openpyxl.load_workbook(patient_excel)
            ws = wb.active
            ws.append([id_val, user_id, full_name, aadhaar_card, email, password, phone_number, gender, dob, address])
            wb.save(patient_excel)

        elif role == 'doctor':
            user_id = generate_user_id("DOC")
            specialization = request.form['specialization']
            hospital_name = request.form['hospital_name']
            cursor.execute("SELECT * FROM doctors WHERE aadhaar_card = %s OR email = %s", (aadhaar_card, email))
            if cursor.fetchone():
                flash("Doctor already registered.", "danger")
                return redirect(url_for('register'))

            cursor.execute('''INSERT INTO doctors (user_id, full_name, aadhaar_card, email, password, phone_number, gender, specialization, hospital_name) 
                              VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                           (user_id, full_name, aadhaar_card, email, password, phone_number, gender, specialization, hospital_name))
            mysql.connection.commit()
            id_val = get_next_id(doctor_excel)
            wb = openpyxl.load_workbook(doctor_excel)
            ws = wb.active
            ws.append([id_val, user_id, full_name, aadhaar_card, email, password, phone_number, gender, specialization, hospital_name])
            wb.save(doctor_excel)

        elif role == 'admin':
            user_id = generate_user_id("ADM")
            hospital_name = request.form['hospital_name']
            cursor.execute("SELECT * FROM admins WHERE aadhaar_card = %s OR email = %s", (aadhaar_card, email))
            if cursor.fetchone():
                flash("Admin already registered.", "danger")
                return redirect(url_for('register'))

            cursor.execute('''INSERT INTO admins (user_id, full_name, aadhaar_card, email, password, phone_number, gender, hospital_name) 
                              VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
                           (user_id, full_name, aadhaar_card, email, password, phone_number, gender, hospital_name))
            mysql.connection.commit()
            id_val = get_next_id(admin_excel)
            wb = openpyxl.load_workbook(admin_excel)
            ws = wb.active
            ws.append([id_val, user_id, full_name, aadhaar_card, email, password, phone_number, gender, hospital_name])
            wb.save(admin_excel)

        flash("Registration successful. Please log in.", "success")
        return redirect(url_for('login'))

    cursor.execute("SELECT name FROM hospitals")
    hospitals = [row[0] for row in cursor.fetchall()]
    return render_template('register.html', hospitals=hospitals)


@app.route('/login', methods=['GET', 'POST']) 
def login():
    if request.method == 'POST':
        aadhaar_or_email = request.form['aadhaar_card']
        password = request.form['password']
        role = request.form['role']
        table = {'patient': 'patients', 'doctor': 'doctors', 'admin': 'admins'}.get(role)

        if not table:
            flash("Invalid role", "danger")
            return redirect(url_for('login'))

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(f"SELECT * FROM {table} WHERE aadhaar_card = %s OR email = %s", (aadhaar_or_email, aadhaar_or_email))
        user = cursor.fetchone()

        if user and user['password'] == password:
            session['loggedin'] = True
            session['role'] = role
            session['user_id'] = user['user_id']
            session['user_name'] = user['full_name']
            session['user'] = user

            if role == 'admin':
                return redirect(url_for('dashboard_admin'))
            elif role == 'doctor':
                return redirect(url_for('dashboard_doctor'))
            else:
                return redirect(url_for('dashboard_patient'))
        else:
            flash("Invalid login credentials", "danger")

    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    role = session.get('role')
    if role == 'admin':
        return redirect(url_for('dashboard_admin'))
    elif role == 'doctor':
        return redirect(url_for('dashboard_doctor'))
    elif role == 'patient':
        return redirect(url_for('dashboard_patient'))
    else:
        return redirect(url_for('login'))

@app.route('/dashboard_admin')
def dashboard_admin():
    if session.get('loggedin') and session.get('role') == 'admin':
        return render_template('dashboard_admin.html', user=session['user'])
    return redirect(url_for('login'))

@app.route('/dashboard_doctor')
def dashboard_doctor():
    if session.get('loggedin') and session.get('role') == 'doctor':
        return render_template('dashboard_doctor.html', user=session['user'])
    return redirect(url_for('login'))

@app.route('/dashboard_patient')
def dashboard_patient():
    if session.get('loggedin') and session.get('role') == 'patient':
        patient_id = session.get("user_id")
        return render_template('dashboard_patient.html',
                               patient=session['user'],
                               appointments=[],
                               doctor={},
                               notifications=[])
    return redirect(url_for('login'))

@app.route('/logout')
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for('login'))

@app.route('/get_patients_with_confirmed_appointments')
def get_patients_with_confirmed_appointments():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT 
            p.user_id AS patient_user_id,
            p.full_name AS patient_name,
            p.phone_number,
            p.dob,
            p.address,
            a.status AS appointment_status,
            a.appointment_time,
            d.user_id AS doctor_user_id,
            d.full_name AS doctor_name,
            d.specialization
        FROM appointments a
        JOIN patients p ON p.user_id = a.patient_id
        JOIN doctors d ON d.user_id = a.doctor_id
        WHERE a.status = 'Confirmed'
    """)
    patients = cursor.fetchall()

    patient_list = []
    for patient in patients:
        patient_list.append({
            "patient_user_id": patient['patient_user_id'],
            "patient_name": patient['patient_name'],
            "phone": patient['phone_number'],
            "dob": str(patient['dob']),
            "address": patient['address'],
            "appointment_status": patient['appointment_status'],
            "appointment_time": str(patient['appointment_time']),  # Ensure time is formatted
            "doctor_user_id": patient['doctor_user_id'],
            "doctor_name": patient['doctor_name'],
            "specialization": patient['specialization']
        })

    return jsonify(patient_list)
@app.route('/admin/add_patient', methods=['POST'])
def admin_add_patient():
    if 'role' not in session or session['role'] != 'admin':
        return jsonify({'error': 'Unauthorized access'}), 403

    full_name = request.form.get('full_name')
    aadhaar_card = request.form.get('aadhaar_card')
    email = request.form.get('email')
    phone_number = request.form.get('phone_number')
    dob = request.form.get('dob')
    address = request.form.get('address')
    gender = request.form.get('gender')
    password = request.form.get('password')

    if not all([full_name, aadhaar_card, email, phone_number, dob, address, gender, password]):
        return jsonify({'error': 'All fields are required'}), 400

    try:
        cursor = mysql.connection.cursor()

        # Check if aadhaar or email already exists
        cursor.execute("SELECT * FROM patients WHERE aadhaar_card=%s OR email=%s", (aadhaar_card, email))
        if cursor.fetchone():
            return jsonify({'error': 'Aadhaar or Email already registered'}), 409

        user_id = "P" + aadhaar_card[-4:]  # Generate user ID

        cursor.execute("""
            INSERT INTO patients (user_id, full_name, aadhaar_card, email, password, phone_number, dob, address, gender, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        """, (user_id, full_name, aadhaar_card, email, password, phone_number, dob, address, gender))
        mysql.connection.commit()

        return jsonify({'success': 'Patient registered successfully'})

    except Exception as e:
        mysql.connection.rollback()
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()



@app.route('/get_doctors', methods=['GET'])
def get_doctors():
    if 'loggedin' not in session or session.get('role') != 'admin':
        return jsonify({"error": "Unauthorized access"}), 401

    admin_hospital = session['user']['hospital_name']  # Get the hospital name of the logged-in admin

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT * FROM doctors WHERE hospital_name = %s", (admin_hospital,))
    doctors = cursor.fetchall()

    if not doctors:
        return jsonify({"message": "No doctors found for this hospital"}), 404

    doctor_list = []
    for doctor in doctors:
        doctor_list.append({
            "doctor_user_id": doctor['user_id'],
            "full_name": doctor['full_name'],
            "specialization": doctor['specialization'],
            "email": doctor['email'],
            "hospital_name": doctor['hospital_name'],
            "phone_number": doctor['phone_number']
        })

    return jsonify(doctor_list)
@app.route('/help_chatbot')
def help_chatbot():
    return render_template('help_chatbot.html')

@app.route('/pharmacy', methods=['GET'])
def pharmacy():
    if 'user_id' not in session:
        flash('Please login to access pharmacy.')
        return redirect(url_for('login'))

    # Fetch hospital list from database
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT name FROM hospitals")
    hospitals = cursor.fetchall()

    return render_template('pharmacy.html', hospitals=hospitals)


@app.route('/upload_pharmacy', methods=['POST'])
def upload_pharmacy():
    patient_id = request.form.get('patient_id')
    patient_name = request.form.get('patient_name')
    hospital_name = request.form.get('hospital_name')
    report_file = request.files.get('report_file')

    if not patient_id or not patient_name or not hospital_name or not report_file:
        flash("All fields are required!")
        return redirect('/pharmacy')

    # Generate order number
    order_number = ''.join(random.choices(string.ascii_uppercase + string.digits, k=12))

    # Extract filename and content type
    filename = report_file.filename
    file_type = report_file.content_type
    file_data = report_file.read()  # Read binary for storing in BLOB

    # Insert into DB with full column list
    cursor = mysql.connection.cursor()
    query = """
        INSERT INTO pharmacy_orders (
            order_number, patient_id, patient_name, hospital_name,
            file_data, filename, file_type
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """
    cursor.execute(query, (
        order_number, patient_id, patient_name, hospital_name,
        file_data, filename, file_type
    ))
    mysql.connection.commit()

    # Fetch hospitals again
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT name FROM hospitals")
    hospitals = cursor.fetchall()

    flash("Report uploaded successfully!")
    return render_template('pharmacy.html', order_number=order_number, hospitals=hospitals)

@app.route('/get_pharmacy_files')
def get_pharmacy_files():
    try:
        admin_hospital = session['user']['hospital_name']  # Get the hospital name of the logged-in admin

        
        print("Hospital in session:", admin_hospital)

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("""
            SELECT order_number, patient_name, filename, file_type, file_data 
            FROM pharmacy_orders 
            WHERE hospital_name = %s
        """, (admin_hospital,))
        results = cursor.fetchall()
        print("Fetched pharmacy files:", results)

        files_data = []
        for row in results:
            if row['file_data'] is None:
                print("Warning: file_data is missing for:", row)
                continue

            file_base64 = base64.b64encode(row['file_data']).decode('utf-8')
            file_url = f"data:{row['file_type']};base64,{file_base64}"

            files_data.append({
                'order_number': row['order_number'],
                'patient_name': row['patient_name'],
                'file_name': row['filename'],
                'file_url': file_url
            })

        cursor.close()
        return jsonify(files_data)

    except Exception as e:
        print("Error fetching pharmacy files:", e)
        return jsonify({'error': 'Failed to fetch pharmacy files'}), 500



@app.route('/patient/profile')
def patient_profile():
    if 'loggedin' in session and session.get('role') == 'patient':
        aadhaar_card = session['user'].get('aadhaar_card')

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("SELECT * FROM patients WHERE aadhaar_card = %s", (aadhaar_card,))
        patient = cursor.fetchone()

        if patient:
            return jsonify(patient)
        else:
            return jsonify({'error': 'Patient not found'}), 404
    else:
        return jsonify({'error': 'Unauthorized'}), 401
@app.route('/tariffdetail')
def tariffdetail():
    return render_template('tariffdetail.html')
@app.route('/abhaa')
def abhaa():
    return render_template('abhaa.html')
@app.route('/labenquiry')
def labenquiry():
    return render_template('labenquiry.html')
@app.route('/hospitalinfo')
def hospitalinfo():
    return render_template('hospitalinfo.html')
@app.route('/billing')
def billing():
    return render_template('billing.html')
@app.route('/opd')
def opd():
    return render_template('opd.html')
# === Route to show appointment form ===
@app.route('/book_appointment', methods=['GET', 'POST'])
def book_appointment():
    if request.method == 'POST':
        # Handle form submission
        patient_id = request.form['patient_id']
        hospital_name = request.form['hospital_name']
        specialization = request.form['specialization']
        doctor_id = request.form['doctor_id']
        appointment_time = request.form['appointment_time']

        cursor = mysql.connection.cursor()
        cursor.execute("""
            INSERT INTO appointments (patient_id, doctor_id, appointment_time, status, hospital_name)
            VALUES (%s, %s, %s, %s, %s)
        """, (patient_id, doctor_id, appointment_time, 'Pending', hospital_name))
        mysql.connection.commit()
        cursor.close()
        return redirect(url_for('book_appointment'))

    # GET: render the form
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT name FROM hospitals")
    hospitals = [row[0] for row in cursor.fetchall()]
    cursor.close()
    return render_template('bookappointments.html', hospitals=hospitals)

# === API: Get specializations for a selected hospital ===
@app.route('/get_specializations/<hospital>')
def get_specializations(hospital):
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT DISTINCT specialization FROM doctors WHERE hospital_name = %s", (hospital,))
    specializations = [row[0] for row in cursor.fetchall()]
    cursor.close()
    return jsonify({'specializations': specializations})

# === API: Get doctors for a hospital + specialization ===
@app.route('/get_doctors/<hospital>/<specialization>')
def get_doctors_filtered(hospital, specialization):
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT user_id, full_name 
        FROM doctors 
        WHERE hospital_name = %s AND specialization = %s
    """, (hospital, specialization))
    doctors = cursor.fetchall()
    cursor.close()
    return jsonify({'doctors': doctors})

@app.route('/manageappointments')
def manage_appointments():
    if 'loggedin' not in session or session.get('role') != 'admin':
        return redirect('/login')

    admin_hospital = session['user']['hospital_name']
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    query = """
        SELECT 
            a.id AS id,
            a.patient_id,
            p.full_name AS patient_name,
            p.phone_number,
            d.full_name AS doctor_name,
            a.appointment_time,
            a.status
        FROM appointments a
        JOIN patients p ON a.patient_id = p.user_id
        JOIN doctors d ON a.doctor_id = d.user_id
        WHERE a.hospital_name = %s AND a.status = 'Pending'
        ORDER BY a.appointment_time DESC
    """
    cursor.execute(query, (admin_hospital,))
    appointments = cursor.fetchall()

    return render_template('manageappointments.html', appointments=appointments)



@app.route('/update_appointment_status/<int:appointment_id>', methods=['POST'])
def update_appointment_status(appointment_id):
    if 'loggedin' not in session or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.get_json()
        new_status = data.get('status')

        if not new_status:
            return jsonify({'error': 'Missing status'}), 400

        # Update database
        cursor = mysql.connection.cursor()
        cursor.execute("UPDATE appointments SET status = %s WHERE id = %s", (new_status, appointment_id))
        mysql.connection.commit()
        cursor.close()

        return jsonify({'message': 'Status updated'}), 200

    except Exception as e:
        print("Error updating appointment:", e)
        return jsonify({'error': 'Something went wrong'}), 500

@app.route('/patient/labreports')
def patient_labreports():
    if 'loggedin' not in session or session['role'] != 'patient':
        return redirect('/login')

    patient_id = session['user']['user_id']

    cursor = mysql.connection.cursor()
    cursor.execute("""
        SELECT report_number, filename, file_type, uploaded_at 
        FROM lab_reports 
        WHERE patient_id = %s 
        ORDER BY uploaded_at DESC
    """, (patient_id,))
    reports = cursor.fetchall()
    cursor.close()

    # Format result as JSON
    result = []
    for row in reports:
        result.append({
            'report_number': row[0],
            'filename': row[1],
            'file_type': row[2],
            'uploaded_at': row[3].strftime('%Y-%m-%d %H:%M')
        })

    return jsonify(result)

@app.route('/download_report/<report_number>')
def download_report(report_number):
    cursor = mysql.connection.cursor()
    cursor.execute("""
        SELECT filename, file_data, file_type 
        FROM lab_reports 
        WHERE report_number = %s
    """, (report_number,))
    report = cursor.fetchone()
    cursor.close()

    if report:
        filename, file_data, file_type = report
        return Response(
            file_data,
            mimetype=file_type,
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    else:
        return "Report not found", 404

@app.route('/patient/appointments')
def patient_appointments():
    if 'loggedin' not in session or session['role'] != 'patient':
        return redirect('/login')

    patient_id = session['user']['user_id']
    print("Patient ID:", patient_id)

    cursor = mysql.connection.cursor()
    cursor.execute("""
        SELECT a.id, d.full_name AS doctor_name, a.appointment_time, a.status, d.specialization, a.hospital_name
        FROM appointments a
        JOIN doctors d ON a.doctor_id = d.user_id
        WHERE a.patient_id = %s
        ORDER BY a.appointment_time DESC
    """, (patient_id,))
    appointments = cursor.fetchall()
    cursor.close()

    result = []
    for row in appointments:
        result.append({
            'appointment_id': row[0],
            'doctor_name': row[1],
            'date': row[2].strftime('%Y-%m-%d'),
            'time': row[2].strftime('%H:%M'),
            'status': row[3],
            'specialization': row[4],
            'hospital_name': row[5]
        })

    return jsonify(result)

from datetime import datetime

@app.route('/patient/reschedule/<appointment_id>', methods=['POST'])
def reschedule_appointment(appointment_id):
    if 'loggedin' not in session or session['role'] != 'patient':
        return jsonify({'error': 'Unauthorized'}), 401

    new_date = request.form.get('date')  # format: YYYY-MM-DD
    new_time = request.form.get('time')  # format: HH:MM

    try:
        new_datetime = datetime.strptime(f"{new_date} {new_time}", "%Y-%m-%d %H:%M")
    except ValueError:
        return jsonify({'error': 'Invalid date/time format'}), 400

    cursor = mysql.connection.cursor()
    cursor.execute("""
        UPDATE appointments
        SET appointment_time = %s, status = 'Pending'
        WHERE id = %s AND patient_id = %s
    """, (new_datetime, appointment_id, session['user']['user_id']))
    mysql.connection.commit()
    cursor.close()

    return jsonify({'success': 'Appointment rescheduled'})
@app.route('/patient/cancel/<appointment_id>', methods=['POST'])
def cancel_appointment(appointment_id):
    try:
        cursor = mysql.connection.cursor()
        cursor.execute("UPDATE appointments SET status = %s WHERE id = %s", ('Cancelled', appointment_id))
        mysql.connection.commit()
        return jsonify({'success': 'Appointment cancelled successfully.'})
    except Exception as e:
        print(e)
        return jsonify({'error': 'Failed to cancel appointment.'})
import qrcode
import io
import base64
import barcode
from barcode.writer import ImageWriter
# Route to show patient ID card
@app.route('/patient/id_card')
def patient_id_card():
    user_id = session.get('user_id')
    if not user_id:
        return redirect('/login')

    # Use DictCursor instead of dictionary=True
    cursor = mysql.connection.cursor(DictCursor)
    cursor.execute("SELECT * FROM patients WHERE user_id = %s", (user_id,))
    patient = cursor.fetchone()
    cursor.close()

    if not patient:
        return "Patient not found", 404

    # Prepare QR data
    patient_info = (
        f"User ID: {patient['user_id']}\n"
        f"Full Name: {patient['full_name']}\n"
        f"Aadhaar: {patient['aadhaar_card']}\n"
        f"Email: {patient['email']}\n"
        f"Phone: {patient['phone_number']}\n"
        f"DOB: {patient['dob']}\n"
        f"Address: {patient['address']}\n"
        f"Gender: {patient['gender']}"
    )

    # QR Code
    qr = qrcode.QRCode(box_size=6, border=2)
    qr.add_data(patient_info)
    qr.make(fit=True)
    img_qr = qr.make_image(fill_color="black", back_color="white")
    buffered_qr = io.BytesIO()
    img_qr.save(buffered_qr, format="PNG")
    qr_base64 = base64.b64encode(buffered_qr.getvalue()).decode()

    # Barcode
    CODE128 = barcode.get_barcode_class('code128')
    barcode_img = CODE128(patient['user_id'], writer=ImageWriter())

    buffered_barcode = io.BytesIO()
    barcode_img.write(buffered_barcode)
    barcode_base64 = base64.b64encode(buffered_barcode.getvalue()).decode()

    return render_template('patient_id_card.html',
                           patient=patient,
                           qr_code=qr_base64,
                           barcode=barcode_base64)
if __name__ == "__main__":
    app.run(debug=True)
